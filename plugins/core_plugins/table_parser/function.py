import csv
from pathlib import Path
from typing import Any, Dict, List, Optional


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_MAX_ROWS = 100


def _resolve(path_str: str) -> Path:
    path = Path(path_str).expanduser()
    if not path.is_absolute():
        path = (PROJECT_ROOT / path).resolve()
    else:
        path = path.resolve()
    try:
        path.relative_to(PROJECT_ROOT)
    except ValueError as exc:
        raise ValueError(f"File '{path}' is outside the project root.") from exc
    return path


def _normalize_rows(raw_rows: List[List[Any]], has_header: bool) -> Dict[str, Any]:
    columns: List[str]
    data_rows: List[List[Any]] = raw_rows[:]
    if has_header and raw_rows:
        columns = [str(col) for col in raw_rows[0]]
        data_rows = raw_rows[1:]
    else:
        max_len = max((len(row) for row in raw_rows), default=0)
        columns = [f"column_{idx+1}" for idx in range(max_len)]

    normalized: List[Dict[str, Any]] = []
    for row in data_rows:
        entry = {}
        for idx, column in enumerate(columns):
            entry[column] = row[idx] if idx < len(row) else None
        normalized.append(entry)

    return {"columns": columns, "rows": normalized}


def _parse_csv(path: Path, has_header: bool, limit: int) -> Dict[str, Any]:
    raw_rows: List[List[Any]] = []
    with path.open(newline="", encoding="utf-8") as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            raw_rows.append(row)
            if len(raw_rows) >= limit + (1 if has_header else 0):
                break
    return _normalize_rows(raw_rows, has_header)


def _parse_excel(path: Path, sheet_name: Optional[str], has_header: bool, limit: int) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ImportError("Reading Excel files requires the 'openpyxl' package.") from exc

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        raw_rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            raw_rows.append(list(row))
            if len(raw_rows) >= limit + (1 if has_header else 0):
                break
    finally:
        wb.close()
    return _normalize_rows(raw_rows, has_header)


def table_parser(
    file_path: str,
    sheet_name: Optional[str] = None,
    max_rows: Optional[int] = None,
    has_header: bool = True,
) -> Dict[str, Any]:
    path = _resolve(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    limit = DEFAULT_MAX_ROWS if max_rows is None else max(1, max_rows)

    suffix = path.suffix.lower()
    if suffix == ".csv":
        parsed = _parse_csv(path, has_header, limit)
    elif suffix in {".xlsx", ".xlsm"}:
        parsed = _parse_excel(path, sheet_name, has_header, limit)
    else:
        raise ValueError("Only CSV and XLSX/XLSM files are supported.")

    parsed["file"] = str(path)
    parsed["row_limit"] = limit
    return parsed
